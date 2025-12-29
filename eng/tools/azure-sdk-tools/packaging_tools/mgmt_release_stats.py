import logging
import pathlib
import os
import re
import datetime
from typing import Tuple
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference


_LOGGER = logging.getLogger(__name__)

FOLDER_SDK = os.path.join(pathlib.Path(__file__).parent.absolute(), "..", "..", "..", "sdk")
SHEET_PACKAGE_MATRIX = "Release Matrix"
SHEET_BREAKING_MATRIX = "Breaking Matrix"
SHEET_EVENT = "Release Log"
SHEET_PIVOT = "Pivot Table"


class ReleaseStats:
    def __init__(self):
        self.wb = Workbook()
        self.wb.active.title = SHEET_EVENT
        self.wb.create_sheet(SHEET_PACKAGE_MATRIX)
        self.wb.create_sheet(SHEET_BREAKING_MATRIX)

        self.date_num = 0
        self.provider_num = 0
        self.release_num = 0
        self.wb.active.append(
            [
                "Release Date",
                "Package Name",
                "Has Breaking(s)",
            ]
        )
        self.sum_sheets = []

    def save(self, xls_file):
        logging.info("Saving result to " + xls_file)
        self.wb.save(xls_file)

    def add_provider(self, package_name, changelog_filename, start_date):
        logging.info("Found provider {}".format(package_name))
        self.provider_num += 1
        self._set_cell(1, self.provider_num + 1, package_name)  # set column title

        with open(changelog_filename) as f:
            in_stable_version = False
            rls_date, rls_version = None, None
            for line in f.readlines():
                version, date = ReleaseStats.parse_release_event(line)
                if version:
                    rls_date, rls_version = date, version
                    if re.match(r"^\d+\.\d+\.\d+$", version):
                        in_stable_version = True
                    else:
                        in_stable_version = False
                    if rls_date >= start_date and in_stable_version:
                        self._add_package_release(self.provider_num + 1, version, rls_date, package_name)

                if rls_date and rls_date >= start_date and ReleaseStats.parse_breaking_head(line):
                    if in_stable_version:
                        self._add_breaking_release(self.provider_num + 1, rls_version, rls_date, package_name)
                    else:
                        logging.info("skip breaking for {}".format(rls_version))

    def gen_pivot_table(self):
        logging.info("Generating pivot table")
        ws_all: Worksheet = self.wb[SHEET_PACKAGE_MATRIX]
        ws_breaking: Worksheet = self.wb[SHEET_BREAKING_MATRIX]
        ws_pivot: Worksheet = self.wb.copy_worksheet(ws_all)
        ws_pivot.title = SHEET_PIVOT
        ws_pivot.cell(1, 1, value="Date")

        # copy breaking values
        col = 2
        while ws_breaking.cell(1, col).value:
            inserted_col = (col - 1) * 2 + 1
            ws_pivot.insert_cols(inserted_col)
            for idx, cell in enumerate(
                ws_breaking["{col_letter}:{col_letter}".format(col_letter=get_column_letter(col))]
            ):
                if idx > 0:
                    ws_pivot.cell(
                        ReleaseStats._find_row_by_date(ws_pivot, ws_breaking.cell(cell.row, 1).value),
                        inserted_col,
                        value=cell.value,
                    )
                else:
                    ws_pivot.cell(cell.row, inserted_col, value="breaking " + cell.value)
            col += 1

        # remove columns who aren't included in the summary sheets
        sum_rps = set()
        for sum_sheet in self.sum_sheets:
            ws = self.wb[sum_sheet]
            col = 2
            while ws.cell(1, col).value:
                sum_rps.add(ws.cell(1, col).value)
                col += 1
        col = 2
        while ws_pivot.cell(1, col).value:
            if ws_pivot.cell(1, col).value in sum_rps:
                col += 2
            else:
                ws_pivot.delete_cols(col, 2)

    @staticmethod
    def _find_row_by_date(ws: Worksheet, d: datetime.date):
        row = 2
        while ws.cell(row, 1).value not in [d, None]:
            row += 1
        return row if ws.cell(row, 1).value == d else None

    def gen_latest_sum(self, days, matrix_sheet_name, threshold):
        logging.info("Generating accumulate sum in {} days with threshold {}".format(days, threshold))
        ws_matrix = self.wb[matrix_sheet_name]
        ws = self.wb.copy_worksheet(ws_matrix)
        ws.title = "{}day>={} {}".format(days, threshold, matrix_sheet_name)
        self.sum_sheets.append(ws.title)

        # add current date in the last row
        last_row, last_date = ReleaseStats._get_last_cell(ws)
        current_date = datetime.datetime.now().date()
        if last_date != current_date:
            last_row += 1
            ws.cell(last_row, 1, value=current_date)

        # accumulate on latest days
        col = 2
        while ws.cell(1, col).value:
            s = 0
            from_row, to_row = 2, 2
            while ws.cell(to_row, 1).value:
                cur_date: datetime.date = ws.cell(to_row, 1).value
                cur_value = ws.cell(to_row, col).value
                if cur_value:
                    s += cur_value
                while (cur_date - ws.cell(from_row, 1).value).days > days:
                    from_value = ws_matrix.cell(from_row, col).value
                    if from_value:
                        s -= from_value
                    from_row += 1
                ws.cell(to_row, col, value=s)
                to_row += 1
            col += 1

        # remove columns which don't meet threshold at the last row.
        col = 2
        while ws.cell(1, col).value:
            if ws.cell(last_row, col).value < threshold:
                ws.delete_cols(col)
            else:
                col += 1
        col = col - 1

        # create chart
        if col > 1:
            values = Reference(ws, min_col=2, min_row=1, max_col=col, max_row=last_row)
            chart = LineChart()
            chart.add_data(values, from_rows=False, titles_from_data=True)
            labels = Reference(ws, min_col=1, min_row=2, max_row=last_row)
            chart.set_categories(labels)
            ws.add_chart(chart)

    @staticmethod
    def _get_last_cell(ws: Worksheet, row=-1, col=1) -> Tuple[int, datetime.date]:
        row_count = 1
        while ws.cell(row_count + 1, col).value:
            row_count += 1
        row = max(row_count + row + 1, 1)
        return row, ws.cell(row, col)

    def _add_package_release(self, col, version, rls_date, package_name):
        logging.debug("Release found: {} {} {}".format(package_name, version, rls_date))

        ws = self.wb[SHEET_PACKAGE_MATRIX]
        row = self.find_or_create_row(rls_date, ws)
        ws.cell(row=row, column=col, value=1)

        ws = self.wb[SHEET_BREAKING_MATRIX]
        row = self.find_or_create_row(rls_date, ws)
        ws.cell(row=row, column=col, value=0)

        ws = self.wb[SHEET_EVENT]
        self.release_num += 1
        ws.append([rls_date, package_name])

    def _add_breaking_release(self, col, version, rls_date, package_name):
        logging.debug("Breaking found: {} {} {}".format(package_name, version, rls_date))

        ws = self.wb[SHEET_BREAKING_MATRIX]
        row = self.find_or_create_row(rls_date, ws)
        ws.cell(row=row, column=col, value=1)

        ws = self.wb[SHEET_EVENT]
        row = self.find_row(ws, rls_date, package_name)
        if row:
            ws.cell(row, 3, 1)

    @staticmethod
    def parse_release_event(line):
        m = re.match(r"^## (\S+) \((\d{4,4})\-(\d{2,2})\-(\d{2,2})\)$", line.strip())
        if not m:
            return None, None
        version, year, month, day = m.groups()
        return version, datetime.date(int(year), int(month), int(day))

    @staticmethod
    def parse_breaking_head(line):
        m = re.match(r"^\*\*.*breaking changes\*\*$", line.strip(), flags=re.IGNORECASE)
        return m is not None

    def find_or_create_row(self, rls_date: datetime.date, ws: Worksheet):
        row = 1
        while True:
            row += 1
            date = ws.cell(row, 1).value
            if date == rls_date:
                return row
            if not date or date > rls_date:
                ws.insert_rows(row)
                ws.cell(row, 1, rls_date)
                return row

    def find_row(self, ws: Worksheet, *values):
        row = 1
        while True:
            row += 1
            match = True
            all_none = True
            for idx, value in enumerate(values):
                v = ws.cell(row, idx + 1).value
                if v:
                    all_none = False
                if v != value:
                    match = False
                    break
            if match:
                return row
            if all_none:
                return None

    def _set_cell(self, row, col, value, sheet_names=[SHEET_PACKAGE_MATRIX, SHEET_BREAKING_MATRIX]):
        for sheet_name in sheet_names:
            sheet = self.wb[sheet_name]
            sheet.cell(row=row, column=col, value=value)

    @staticmethod
    def _get_column_count(ws: Worksheet):
        col = 1
        while ws.cell(1, col + 1).value:
            col += 1
        return col

    def sum_matrix_sheets(self):
        logging.info("Adding Sum and Total to Matix")
        for sheet_name in [SHEET_PACKAGE_MATRIX, SHEET_BREAKING_MATRIX]:
            ws = self.wb[sheet_name]
            last_row, last_date = ReleaseStats._get_last_cell(ws)
            last_col = ReleaseStats._get_column_count(ws)

            last_row, last_col = last_row + 1, last_col + 1
            ws.cell(1, last_col, value="Sum")
            for r in range(2, last_row):
                ws.cell(
                    r,
                    last_col,
                    value="=sum(B{row}:{col}{row})".format(row=r, col=get_column_letter(last_col - 1)),
                )

            ws.cell(last_row, 1, value="Total")
            for c in range(2, last_col + 1):
                ws.cell(
                    last_row,
                    c,
                    value="=sum({col}2:{col}{row})".format(row=last_row - 1, col=get_column_letter(c)),
                )


def build_release_stats(xls_file, threshold, start_date):
    stats = ReleaseStats()
    for root, dirs, files in os.walk(FOLDER_SDK):
        for file in files:
            if file == "CHANGELOG.md":
                package_name = root.split(os.sep)[-1]
                if not package_name.startswith("azure-mgmt-"):
                    continue
                stats.add_provider(package_name, os.path.join(root, file), start_date)
    stats.gen_latest_sum(90, SHEET_BREAKING_MATRIX, threshold - 1)
    stats.gen_latest_sum(180, SHEET_BREAKING_MATRIX, threshold)
    stats.gen_latest_sum(90, SHEET_PACKAGE_MATRIX, threshold - 1)
    stats.gen_pivot_table()
    stats.sum_matrix_sheets()
    stats.save(xls_file)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Release history statistics",
    )
    parser.add_argument("--debug", help="Verbosity in DEBUG mode")
    parser.add_argument("--xlsx", default="release-stats.xlsx", help="Output xlsx filename")
    parser.add_argument(
        "--threshold",
        default=3,
        type=int,
        help="Only stats RPs has breakings above this number recently",
    )
    parser.add_argument(
        "--start-date",
        default=datetime.date(2010, 1, 1),
        type=lambda d: datetime.datetime.strptime(d, "%Y%m%d").date(),
        help="Start from date in format yyyymmdd",
    )

    args = parser.parse_args()

    logging.basicConfig(level=logging.DEBUG if args.debug else logging.INFO)

    build_release_stats(args.xlsx, args.threshold, args.start_date)
